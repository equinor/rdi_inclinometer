'''
Created on 2. des. 2014
@author: PPAR
'''
import os
import openpyxl

class ExcelManager(object):
    
    def __init__(self, working_directory, working_file_name, sheet_name):
        self.working_directory = working_directory
        self.working_file_name = working_file_name
        self.file_name = working_directory + working_file_name + ".xlsx"
        self.sheet_name = sheet_name
    


    def _find_sheet(self):
        self.sheet = self.workbook.get_sheet_by_name(self.sheet_name)
        #Create the sheet if it does not exist
        if (self.sheet == None):
            self.sheet = self.workbook.create_sheet(0, self.sheet_name)

    def _open_excel_file_and_relevant_sheet(self):
        if (os.path.isfile(self.file_name)):
            self.workbook = openpyxl.load_workbook(filename=self.file_name)
            self._find_sheet()
        else:
            self.workbook = openpyxl.Workbook()
            self._find_sheet()
            #Add headers
            self.sheet.append(["ID","Timestamp", "Data Type", "AccX (g)", "AccY (g)", "AccZ (g)", "AngX (degrees per sec)", "AngY (degrees per sec)", "AngZ (degrees per sec)", "MagX (Gauss)", "MagY (Gauss)", "MagZ (Gauss)"])

    def _save_and_close_excel_file(self):
        self.workbook.save(self.file_name)
        
    def write_data(self, timestamp, data_type, acc_data = [], gyro_data = [], magn_data= []):
        self._open_excel_file_and_relevant_sheet()
        
        max_row = self.sheet.get_highest_row()
       
        row = []
        row.append(max_row + 1)
        row.append(timestamp)
        row.append(data_type)
        for i in acc_data:
            row.append(i)
        
        for i in gyro_data:
            row.append(i)
        
        for i in magn_data:
            row.append(i)
        
        self.sheet.append(row)
        
        
        self._save_and_close_excel_file()